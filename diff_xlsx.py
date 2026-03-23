#!/usr/bin/env python3
"""
Diff Tool
---------
Compares two .xlsx or two .csv files with the same format and generates an HTML diff report.

Usage:
    python diff_xlsx.py
"""

import os
import sys
import csv
import glob
import re
import shutil
from openpyxl import load_workbook
from collections import OrderedDict


# ─── File Discovery ───────────────────────────────────────────────────────────

def get_files():
    """Return (file1_path, file2_path) from the 'source' directory (xlsx or csv, same format)."""
    source_dir = os.path.join(os.getcwd(), "source")
    if not os.path.isdir(source_dir):
        raise FileNotFoundError(f"'source' directory not found at: {os.getcwd()}")

    files = sorted(
        glob.glob(os.path.join(source_dir, "*.xlsx")) +
        glob.glob(os.path.join(source_dir, "*.csv"))
    )

    if len(files) == 0:
        raise FileNotFoundError("No .xlsx or .csv files found in 'source' directory.")
    if len(files) == 1:
        raise FileNotFoundError("Only 1 file found in 'source'; need exactly 2.")

    def same_format(f1, f2):
        return os.path.splitext(f1)[1].lower() == os.path.splitext(f2)[1].lower()

    if len(files) == 2:
        if not same_format(files[0], files[1]):
            raise ValueError(
                f"The two files have different formats "
                f"({os.path.splitext(files[0])[1]} vs {os.path.splitext(files[1])[1]}). "
                f"Both files must be the same format (.xlsx or .csv)."
            )
        return files[0], files[1]

    # More than 2: let the user pick, enforcing same format
    print(f"\nFound {len(files)} files in 'source':")
    for i, f in enumerate(files, 1):
        print(f"  {i}. {os.path.basename(f)}")

    while True:
        try:
            a = int(input("\nEnter the number of FILE 1: ").strip())
            b = int(input("Enter the number of FILE 2: ").strip())
            if a == b:
                print("Please choose two different files.")
                continue
            if not (1 <= a <= len(files) and 1 <= b <= len(files)):
                print(f"Please enter numbers between 1 and {len(files)}.")
                continue
            f1, f2 = files[a - 1], files[b - 1]
            if not same_format(f1, f2):
                ext1 = os.path.splitext(f1)[1]
                ext2 = os.path.splitext(f2)[1]
                print(
                    f"Format mismatch: '{os.path.basename(f1)}' is {ext1} and "
                    f"'{os.path.basename(f2)}' is {ext2}. Please choose two files of the same format."
                )
                continue
            return f1, f2
        except ValueError:
            print("Please enter a valid number.")


# ─── Reading ──────────────────────────────────────────────────────────────────

def is_blank(v):
    return v is None or str(v).strip() == ""


def read_xlsx(filepath):
    """
    Read the first sheet of an xlsx file.
    Returns (headers, data_rows) after:
      - Stripping trailing empty rows
      - Removing fully-empty columns (every cell, including header, is blank)
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    raw = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    if not raw:
        return [], []

    # Strip trailing empty rows
    while raw and all(is_blank(c) for c in raw[-1]):
        raw.pop()
    if not raw:
        return [], []

    # Pad all rows to the same width
    width = max(len(r) for r in raw)
    for r in raw:
        r += [None] * (width - len(r))

    # Identify non-empty columns (at least one non-blank value across ALL rows)
    keep = [
        col for col in range(width)
        if any(not is_blank(raw[row][col]) for row in range(len(raw)))
    ]

    headers = [raw[0][c] for c in keep]
    data_rows = [[raw[r][c] for c in keep] for r in range(1, len(raw))]

    # Strip trailing empty data rows once more (after column filtering)
    while data_rows and all(is_blank(c) for c in data_rows[-1]):
        data_rows.pop()

    return headers, data_rows


def detect_delimiter(filepath):
    """Sniff the delimiter from the first 4 KB of a CSV file."""
    with open(filepath, encoding="utf-8", newline="") as f:
        sample = f.read(4096)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","  # fallback


def read_csv(filepath):
    """
    Read a CSV file, auto-detecting the delimiter.
    Returns (headers, data_rows) after:
      - Stripping trailing empty rows
      - Removing fully-empty columns (every cell, including header, is blank)
    """
    delimiter = detect_delimiter(filepath)
    with open(filepath, encoding="utf-8", newline="") as f:
        raw = [row for row in csv.reader(f, delimiter=delimiter)]

    if not raw:
        return [], []

    # Strip trailing empty rows
    while raw and all(is_blank(c) for c in raw[-1]):
        raw.pop()
    if not raw:
        return [], []

    # Pad all rows to the same width
    width = max(len(r) for r in raw)
    for r in raw:
        r += [""] * (width - len(r))

    # Identify non-empty columns (at least one non-blank value across ALL rows)
    keep = [
        col for col in range(width)
        if any(not is_blank(raw[row][col]) for row in range(len(raw)))
    ]

    headers = [raw[0][c] for c in keep]
    data_rows = [[raw[r][c] for c in keep] for r in range(1, len(raw))]

    # Strip trailing empty data rows once more (after column filtering)
    while data_rows and all(is_blank(c) for c in data_rows[-1]):
        data_rows.pop()

    return headers, data_rows


def read_sheet(filepath):
    """Dispatch to the appropriate reader based on file extension."""
    if os.path.splitext(filepath)[1].lower() == ".csv":
        return read_csv(filepath)
    return read_xlsx(filepath)


# ─── Format Check ─────────────────────────────────────────────────────────────

def normalize_header(h):
    return str(h).strip() if h is not None else ""


def check_format(headers1, headers2):
    """Raise ValueError if column headers differ."""
    n1 = [normalize_header(h) for h in headers1]
    n2 = [normalize_header(h) for h in headers2]
    if n1 != n2:
        raise ValueError(
            f"Files have different column headers!\n"
            f"  File 1 ({len(n1)} cols): {n1}\n"
            f"  File 2 ({len(n2)} cols): {n2}"
        )


# ─── Aliases ──────────────────────────────────────────────────────────────────

def load_aliases():
    """
    Load column-specific value aliases from aliases.txt in the working directory.
    Format (one per line):  ColumnName:(value1,value2)
    Returns dict: { column_name: (value1, value2) }
    """
    path = os.path.join(os.getcwd(), "aliases.txt")
    if not os.path.isfile(path):
        input("\nNo aliases.txt found. Press Enter to continue without aliases...")
        return {}

    aliases = {}
    print("\nLoading aliases.txt...")
    with open(path, encoding="utf-8") as f:
        for lineno, line in enumerate(f, 1):
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            match = re.match(r'^(.+?):\((.+?),(.+?)\)$', line)
            if not match:
                print(f"  Warning: skipping malformed line {lineno}: {line!r}")
                continue
            col = match.group(1).strip()
            v1  = match.group(2).strip()
            v2  = match.group(3).strip()
            if col not in aliases:
                aliases[col] = []
            aliases[col].append((v1, v2))
    print(f"  Loaded {len(aliases)} alias(es) for column(s): {list(aliases.keys())}")
    return aliases


# ─── Ignore Substrings ────────────────────────────────────────────────────────

def load_ignore_substrings():
    """
    Load substring substitution rules from ignore_substring.txt in the working directory.
    Format (one per line):  filename:column:(find,replacement)
    Returns dict: { filename: { column: [(find, replacement), ...] } }
    Rules for each column are ordered top-to-bottom as written in the file.
    """
    path = os.path.join(os.getcwd(), "ignore_substring.txt")
    if not os.path.isfile(path):
        input("\nNo ignore_substring.txt found. Press Enter to continue without substring rules...")
        return {}

    rules = {}
    print("\nLoading ignore_substring.txt...")
    with open(path, encoding="utf-8") as f:
        for lineno, line in enumerate(f, 1):
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            match = re.match(r'^(.+?):(.+?):\((.+?),(.*?)\)$', line)
            if not match:
                print(f"  Warning: skipping malformed line {lineno}: {line!r}")
                continue
            filename = match.group(1).strip()
            col      = match.group(2).strip()
            find     = match.group(3)
            replace  = match.group(4)
            rules.setdefault(filename, {}).setdefault(col, []).append((find, replace))

    total = sum(len(cols) for cols in rules.values())
    print(f"  Loaded {total} rule(s) across {len(rules)} file(s): {list(rules.keys())}")
    return rules


def apply_substitutions(value, sub_rules):
    """Apply a list of (find, replacement) rules in order to value, replacing only the first occurrence each time."""
    s = str(value) if value is not None else ""
    for find, replacement in sub_rules:
        s = s.replace(find, replacement, 1)
    return s


# ─── Extra Report HTML Generation ────────────────────────────────────────────

def generate_extra_html(diffs, headers, file1_name, file2_name, col_sub1, col_sub2,
                        part=1, total_parts=1, base_filename="diff_report_substituted",
                        diff_col_indices=None):
    """
    Generate an HTML report showing only changed rows with post-substitution values,
    and only the columns that had at least one difference.
    diff_col_indices can be pre-computed from the full dataset to ensure consistent
    column layout across split parts.
    """
    header_strs = [normalize_header(h) for h in headers]

    # Collect all column indices that differ across all changed rows
    if diff_col_indices is None:
        diff_col_indices = sorted({i for d in diffs if d["type"] == "changed" for i in d["changed"]})

    if not diff_col_indices:
        return None

    diff_headers = [header_strs[i] for i in diff_col_indices]

    rows_html = []
    for d in diffs:
        if d["type"] != "changed":
            continue

        # Apply substitutions to get the values the tool actually compared
        r1_sub = [
            apply_substitutions(d["row1"][i], col_sub1[i]) if col_sub1[i] else (d["row1"][i] if d["row1"][i] is not None else "")
            for i in diff_col_indices
        ]
        r2_sub = [
            apply_substitutions(d["row2"][i], col_sub2[i]) if col_sub2[i] else (d["row2"][i] if d["row2"][i] is not None else "")
            for i in diff_col_indices
        ]

        # Map original diff_col_indices to new 0-based positions for highlighting
        changed_new = {diff_col_indices.index(i) for i in d["changed"]}

        rows_html.append(build_data_row(r1_sub, "old", "lbl-old", "OLD", d["label"], changed_new))
        rows_html.append(build_data_row(r2_sub, "new", "lbl-new", "NEW", d["label"], changed_new))

    th_cols = "".join(f"<th>{esc(h)}</th>" for h in diff_headers)
    table_html = (
        f"<table>"
        f"<thead><tr><th>Row / Key</th><th>Type</th>{th_cols}</tr></thead>"
        f"<tbody>{''.join(rows_html)}</tbody>"
        f"</table>"
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Diff Report (Post-Substitution)</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{
    font-family: Arial, sans-serif;
    font-size: 13px;
    margin: 24px;
    background: #f0f2f5;
    color: #222;
  }}
  h1 {{ margin-bottom: 12px; color: #1e293b; }}
  .summary {{
    background: #fff;
    padding: 14px 20px;
    border-radius: 8px;
    margin-bottom: 22px;
    box-shadow: 0 1px 4px rgba(0,0,0,.12);
  }}
  .summary p {{ margin: 4px 0; }}
  table {{
    width: 100%;
    border-collapse: collapse;
    background: #fff;
    box-shadow: 0 1px 4px rgba(0,0,0,.12);
    border-radius: 8px;
    overflow: hidden;
  }}
  th {{
    background: #1e293b;
    color: #f8fafc;
    padding: 9px 12px;
    text-align: left;
    font-size: 12px;
    white-space: nowrap;
  }}
  td {{
    padding: 6px 12px;
    border-bottom: 1px solid #e5e7eb;
    vertical-align: top;
    word-break: break-word;
    max-width: 280px;
  }}
  td.meta {{ white-space: nowrap; font-size: 11px; font-weight: bold; min-width: 90px; }}
  td.lbl-old {{ background: #fef2f2; color: #b91c1c; border-right: 3px solid #fca5a5; }}
  td.lbl-new {{ background: #f0fdf4; color: #15803d; border-right: 3px solid #86efac; }}
  td.old     {{ background: #fef2f2; }}
  td.new     {{ background: #f0fdf4; }}
  td.cell-hl {{ background: #fef08a !important; font-weight: bold; outline: 1px solid #ca8a04; }}
  tr:last-child td {{ border-bottom: none; }}
  .nav {{ margin-bottom: 16px; font-size: 13px; }}
  .nav a {{ color: #1e293b; text-decoration: none; font-weight: bold; }}
  .nav a:hover {{ text-decoration: underline; }}
</style>
</head>
<body>
<h1>Diff Report (Post-Substitution Values)</h1>
<div class="summary">
  <p><strong>File 1:</strong> {esc(file1_name)}</p>
  <p><strong>File 2:</strong> {esc(file2_name)}</p>
  <p>Cell values shown after substitution rules were applied (what the tool actually compared).</p>
  <p>Only changed rows and differing columns are included.</p>
</div>
{_nav_html(part, total_parts, base_filename)}
{table_html}
{_nav_html(part, total_parts, base_filename)}
</body>
</html>"""


# ─── CSV Report Generation ────────────────────────────────────────────────────

def _summary_rows(file1_name, file2_name, case_sensitive, ignore_substrings,
                  total_rows_compared, diffs, skip_columns, headers):
    """Return a list of [label, value] rows for the summary section."""
    header_strs = [normalize_header(h) for h in headers]
    skip_columns = skip_columns or set()
    n_changed = sum(1 for d in diffs if d["type"] == "changed")
    n_added   = sum(1 for d in diffs if d["type"] == "added")
    n_deleted = sum(1 for d in diffs if d["type"] == "deleted")

    col_diff_counts = {h: 0 for h in header_strs if h not in skip_columns}
    for d in diffs:
        if d["type"] == "changed":
            for i in d["changed"]:
                col = header_strs[i]
                if col in col_diff_counts:
                    col_diff_counts[col] += 1

    rows = [
        ["File 1", file1_name],
        ["File 2", file2_name],
        ["Comparison mode", "Case-sensitive" if case_sensitive else "Case-insensitive"],
        ["Total rows compared", total_rows_compared],
        [],
        ["Changed rows", n_changed],
        ["Added rows (only in File 2)", n_added],
        ["Deleted rows (only in File 1)", n_deleted],
        [],
        ["Changed cells per column", ""],
    ]
    for col, count in col_diff_counts.items():
        if count > 0:
            rows.append([col, count])
    rows.append([])
    return rows


def generate_csv_report(diffs, headers, file1_name, file2_name, case_sensitive=True,
                        ignore_substrings=None, total_rows_compared=0, skip_columns=None):
    import io
    header_strs = [normalize_header(h) for h in headers]
    output = io.StringIO()
    writer = csv.writer(output)

    for row in _summary_rows(file1_name, file2_name, case_sensitive, ignore_substrings,
                             total_rows_compared, diffs, skip_columns, headers):
        writer.writerow(row)

    writer.writerow(["diff_type", "row_key"] + header_strs)
    for d in diffs:
        if d["type"] == "changed":
            writer.writerow(["changed_old", d["label"]] + [v if v is not None else "" for v in d["row1"]])
            writer.writerow(["changed_new", d["label"]] + [v if v is not None else "" for v in d["row2"]])
        elif d["type"] == "added":
            writer.writerow(["added", d["label"]] + [v if v is not None else "" for v in d["row2"]])
        elif d["type"] == "deleted":
            writer.writerow(["deleted", d["label"]] + [v if v is not None else "" for v in d["row1"]])

    return output.getvalue()


def generate_extra_csv(diffs, headers, file1_name, file2_name, col_sub1, col_sub2,
                       diff_col_indices=None):
    import io
    header_strs = [normalize_header(h) for h in headers]
    if diff_col_indices is None:
        diff_col_indices = sorted({i for d in diffs if d["type"] == "changed" for i in d["changed"]})
    if not diff_col_indices:
        return None

    diff_headers = [header_strs[i] for i in diff_col_indices]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["File 1", file1_name])
    writer.writerow(["File 2", file2_name])
    writer.writerow(["Note", "Cell values shown after substitution rules applied. Only changed rows and differing columns included."])
    writer.writerow([])
    writer.writerow(["diff_type", "row_key"] + diff_headers)

    for d in diffs:
        if d["type"] != "changed":
            continue
        r1_sub = [
            apply_substitutions(d["row1"][i], col_sub1[i]) if col_sub1[i] else (d["row1"][i] if d["row1"][i] is not None else "")
            for i in diff_col_indices
        ]
        r2_sub = [
            apply_substitutions(d["row2"][i], col_sub2[i]) if col_sub2[i] else (d["row2"][i] if d["row2"][i] is not None else "")
            for i in diff_col_indices
        ]
        writer.writerow(["changed_old", d["label"]] + r1_sub)
        writer.writerow(["changed_new", d["label"]] + r2_sub)

    return output.getvalue()


# ─── XLSX Report Generation ───────────────────────────────────────────────────

def _xlsx_styles():
    from openpyxl.styles import PatternFill, Font, Alignment
    return {
        "header_fill":  PatternFill("solid", fgColor="1E293B"),
        "header_font":  Font(color="F8FAFC", bold=True),
        "old_fill":     PatternFill("solid", fgColor="FEE2E2"),
        "new_fill":     PatternFill("solid", fgColor="F0FDF4"),
        "added_fill":   PatternFill("solid", fgColor="DCFCE7"),
        "deleted_fill": PatternFill("solid", fgColor="FEE2E2"),
        "hl_fill":      PatternFill("solid", fgColor="FEF08A"),
        "label_font":   Font(bold=True),
        "wrap":         Alignment(wrap_text=True, vertical="top"),
    }


def _write_xlsx_summary(ws, summary_rows, styles):
    from openpyxl.styles import Font
    for row in summary_rows:
        if not row:
            ws.append([])
        else:
            ws.append(row)
            ws.cell(ws.max_row, 1).font = Font(bold=True)


def _write_xlsx_diff_rows(ws, diffs, header_strs, styles, col_offset=2):
    """Write diff data rows. col_offset=2 accounts for diff_type and row_key columns."""
    for d in diffs:
        dtype = d["type"]
        if dtype == "changed":
            for side, row_data, row_fill in [("changed_old", d["row1"], styles["old_fill"]),
                                              ("changed_new", d["row2"], styles["new_fill"])]:
                ws.append([side, d["label"]] + [v if v is not None else "" for v in row_data])
                r = ws.max_row
                for c in range(1, len(header_strs) + col_offset + 1):
                    ws.cell(r, c).fill = row_fill
                    ws.cell(r, c).alignment = styles["wrap"]
                for i in d["changed"]:
                    ws.cell(r, col_offset + 1 + i).fill = styles["hl_fill"]
        elif dtype == "added":
            ws.append(["added", d["label"]] + [v if v is not None else "" for v in d["row2"]])
            r = ws.max_row
            for c in range(1, len(header_strs) + col_offset + 1):
                ws.cell(r, c).fill = styles["added_fill"]
                ws.cell(r, c).alignment = styles["wrap"]
        elif dtype == "deleted":
            ws.append(["deleted", d["label"]] + [v if v is not None else "" for v in d["row1"]])
            r = ws.max_row
            for c in range(1, len(header_strs) + col_offset + 1):
                ws.cell(r, c).fill = styles["deleted_fill"]
                ws.cell(r, c).alignment = styles["wrap"]


def _write_xlsx_header_row(ws, col_headers, styles):
    ws.append(col_headers)
    r = ws.max_row
    for c in range(1, len(col_headers) + 1):
        ws.cell(r, c).fill = styles["header_fill"]
        ws.cell(r, c).font = styles["header_font"]


def generate_xlsx_report(diffs, headers, file1_name, file2_name, case_sensitive=True,
                         ignore_substrings=None, total_rows_compared=0, skip_columns=None):
    from openpyxl import Workbook
    header_strs = [normalize_header(h) for h in headers]
    styles = _xlsx_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Diff"

    summary = _summary_rows(file1_name, file2_name, case_sensitive, ignore_substrings,
                            total_rows_compared, diffs, skip_columns, headers)
    _write_xlsx_summary(ws, summary, styles)
    _write_xlsx_header_row(ws, ["diff_type", "row_key"] + header_strs, styles)
    _write_xlsx_diff_rows(ws, diffs, header_strs, styles)
    return wb


def generate_extra_xlsx(diffs, headers, file1_name, file2_name, col_sub1, col_sub2,
                        diff_col_indices=None):
    from openpyxl import Workbook
    header_strs = [normalize_header(h) for h in headers]
    if diff_col_indices is None:
        diff_col_indices = sorted({i for d in diffs if d["type"] == "changed" for i in d["changed"]})
    if not diff_col_indices:
        return None

    diff_headers = [header_strs[i] for i in diff_col_indices]
    styles = _xlsx_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Diff (Post-Substitution)"

    from openpyxl.styles import Font
    for label, value in [("File 1", file1_name), ("File 2", file2_name),
                         ("Note", "Cell values after substitution rules applied. Only changed rows and differing columns.")]:
        ws.append([label, value])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.append([])

    _write_xlsx_header_row(ws, ["diff_type", "row_key"] + diff_headers, styles)

    for d in diffs:
        if d["type"] != "changed":
            continue
        r1_sub = [
            apply_substitutions(d["row1"][i], col_sub1[i]) if col_sub1[i] else (d["row1"][i] if d["row1"][i] is not None else "")
            for i in diff_col_indices
        ]
        r2_sub = [
            apply_substitutions(d["row2"][i], col_sub2[i]) if col_sub2[i] else (d["row2"][i] if d["row2"][i] is not None else "")
            for i in diff_col_indices
        ]
        changed_new = {diff_col_indices.index(i) for i in d["changed"]}

        for side, row_data, row_fill in [("changed_old", r1_sub, styles["old_fill"]),
                                          ("changed_new", r2_sub, styles["new_fill"])]:
            ws.append([side, d["label"]] + row_data)
            r = ws.max_row
            for c in range(1, len(diff_headers) + 3):
                ws.cell(r, c).fill = row_fill
                ws.cell(r, c).alignment = styles["wrap"]
            for ci in changed_new:
                ws.cell(r, 3 + ci).fill = styles["hl_fill"]

    return wb


# ─── User Prompts ─────────────────────────────────────────────────────────────

def ask_unique_key(headers):
    """
    Ask the user whether there is a unique key column.
    Returns (has_key: bool, key_col: str or None).
    """
    answer = input("\nDoes the data have a unique key column? (yes/no) [no]: ").strip().lower()
    if answer not in ("yes", "y"):
        return False, None

    default = normalize_header(headers[0]) if headers else ""
    raw = input(f"Enter the key column name [default: '{default}']: ").strip()
    key_col = raw if raw else default

    header_strs = [normalize_header(h) for h in headers]
    if key_col not in header_strs:
        raise ValueError(
            f"Key column '{key_col}' not found in headers: {header_strs}"
        )
    return True, key_col


def ask_case_sensitive():
    """Ask whether comparison should be case-sensitive. Default: yes (press Enter)."""
    answer = input("\nUse case-sensitive comparison? (yes/no) [yes]: ").strip().lower()
    return answer not in ("no", "n")


def ask_export_format():
    """Ask the user for the export format. Default: html."""
    raw = input("\nExport format (html/csv/xlsx) [html]: ").strip().lower()
    if raw in ("csv", "xlsx"):
        return raw
    return "html"


def ask_split_report():
    """Ask whether to split the report into multiple files. Default: no.
    Returns (should_split: bool, rows_per_file: int or None).
    """
    answer = input("\nSplit report into multiple files? (yes/no) [no]: ").strip().lower()
    if answer not in ("yes", "y"):
        return False, None
    raw = input("Rows per file [1000]: ").strip()
    try:
        rows_per_file = int(raw) if raw else 1000
        if rows_per_file < 1:
            raise ValueError
    except ValueError:
        print("Invalid number, using default 1000.")
        rows_per_file = 1000
    return True, rows_per_file


def ask_extra_report():
    """Ask whether to export an additional report with post-substitution values. Default: no."""
    answer = input("\nExport additional report with post-substitution values? (yes/no) [no]: ").strip().lower()
    return answer in ("yes", "y")


def ask_skip_columns(headers):
    """
    Ask the user for column names to skip during comparison (comma-separated).
    Returns a set of normalized column name strings to ignore.
    """
    raw = input("\nEnter column(s) to skip during comparison (comma-separated, or press Enter to skip none): ").strip()
    if not raw:
        return set()

    header_strs = [normalize_header(h) for h in headers]
    skip = set()
    for name in raw.split(","):
        name = name.strip()
        if name in header_strs:
            skip.add(name)
        else:
            print(f"  Warning: column '{name}' not found in headers, skipping.")
    return skip


# ─── Comparison ───────────────────────────────────────────────────────────────

def _try_numeric(s):
    """Return float if s is a numeric string, else None."""
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def values_differ(a, b, case_sensitive=True):
    """Return True if two cell values should be considered different."""
    if a is None and b is None:
        return False
    # Native numeric comparison for xlsx values (always exact)
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return a != b
    sa = str(a if a is not None else "")
    sb = str(b if b is not None else "")
    if case_sensitive:
        return sa != sb
    # Case-insensitive: normalize numeric-looking strings, then fold case
    na, nb = _try_numeric(sa), _try_numeric(sb)
    if na is not None and nb is not None:
        return na != nb
    return sa.lower() != sb.lower()


def are_aliases(a, b, alias_pairs):
    """Return True if a and b match any defined alias pair for this column (bidirectional, exact match)."""
    sa = str(a if a is not None else "")
    sb = str(b if b is not None else "")
    return any(
        (sa == v1 and sb == v2) or (sa == v2 and sb == v1)
        for v1, v2 in alias_pairs
    )


def changed_indices(row1, row2, case_sensitive=True, col_aliases=None, skip_indices=None,
                    col_sub1=None, col_sub2=None):
    result = []
    for i in range(len(row1)):
        if skip_indices and i in skip_indices:
            continue
        # Apply per-file substring substitutions before comparison
        a = apply_substitutions(row1[i], col_sub1[i]) if col_sub1 and col_sub1[i] else row1[i]
        b = apply_substitutions(row2[i], col_sub2[i]) if col_sub2 and col_sub2[i] else row2[i]
        alias = col_aliases[i] if col_aliases else None
        if alias and are_aliases(a, b, alias):
            continue
        if values_differ(a, b, case_sensitive):
            result.append(i)
    return result


def _build_col_subs(header_strs, ignore_substrings, filename):
    """Return a per-column list of substitution rules for the given filename."""
    file_rules = ignore_substrings.get(filename, {}) if ignore_substrings else {}
    return [file_rules.get(h, []) for h in header_strs]


def compare_by_key(headers, rows1, rows2, key_col, case_sensitive=True, aliases=None,
                   skip_columns=None, ignore_substrings=None, file1_name="", file2_name=""):
    """Compare rows matched by a unique key column."""
    header_strs = [normalize_header(h) for h in headers]
    ki = header_strs.index(key_col)
    col_aliases  = [aliases.get(h) for h in header_strs] if aliases else None
    skip_indices = {i for i, h in enumerate(header_strs) if skip_columns and h in skip_columns}
    col_sub1     = _build_col_subs(header_strs, ignore_substrings, file1_name)
    col_sub2     = _build_col_subs(header_strs, ignore_substrings, file2_name)

    def to_dict(rows):
        d = OrderedDict()
        for r in rows:
            k = str(r[ki]) if r[ki] is not None else ""
            d[k] = r
        return d

    d1, d2 = to_dict(rows1), to_dict(rows2)
    all_keys = list(d1) + [k for k in d2 if k not in d1]

    diffs = []
    for key in all_keys:
        in1, in2 = key in d1, key in d2
        if in1 and not in2:
            diffs.append({
                "type": "deleted", "label": f"Key: {key}",
                "row1": d1[key], "row2": None, "changed": set(),
            })
        elif not in1 and in2:
            diffs.append({
                "type": "added", "label": f"Key: {key}",
                "row1": None, "row2": d2[key], "changed": set(),
            })
        else:
            ci = changed_indices(d1[key], d2[key], case_sensitive, col_aliases,
                                 skip_indices, col_sub1, col_sub2)
            if ci:
                diffs.append({
                    "type": "changed", "label": f"Key: {key}",
                    "row1": d1[key], "row2": d2[key], "changed": set(ci),
                })
    return diffs


def compare_by_position(headers, rows1, rows2, case_sensitive=True, aliases=None,
                        skip_columns=None, ignore_substrings=None, file1_name="", file2_name=""):
    """Compare rows by their position (row index)."""
    header_strs  = [normalize_header(h) for h in headers]
    col_aliases  = [aliases.get(h) for h in header_strs] if aliases else None
    skip_indices = {i for i, h in enumerate(header_strs) if skip_columns and h in skip_columns}
    col_sub1     = _build_col_subs(header_strs, ignore_substrings, file1_name)
    col_sub2     = _build_col_subs(header_strs, ignore_substrings, file2_name)

    diffs = []
    for i in range(max(len(rows1), len(rows2))):
        label = f"Row {i + 2}"  # +2: 1-indexed + skip header row
        r1 = rows1[i] if i < len(rows1) else None
        r2 = rows2[i] if i < len(rows2) else None

        if r1 is None:
            diffs.append({
                "type": "added", "label": label,
                "row1": None, "row2": r2, "changed": set(),
            })
        elif r2 is None:
            diffs.append({
                "type": "deleted", "label": label,
                "row1": r1, "row2": None, "changed": set(),
            })
        else:
            ci = changed_indices(r1, r2, case_sensitive, col_aliases,
                                 skip_indices, col_sub1, col_sub2)
            if ci:
                diffs.append({
                    "type": "changed", "label": label,
                    "row1": r1, "row2": r2, "changed": set(ci),
                })
    return diffs


# ─── HTML Generation ──────────────────────────────────────────────────────────

def esc(text):
    """Escape HTML special characters."""
    s = "" if text is None else str(text)
    return (
        s.replace("&", "&amp;")
         .replace("<", "&lt;")
         .replace(">", "&gt;")
         .replace('"', "&quot;")
    )


def build_data_row(vals, row_cls, label_cls, side_label, label_text, changed_set):
    cells = "".join(
        f'<td class="{row_cls}{"  cell-hl" if i in changed_set else ""}">{esc(v)}</td>'
        for i, v in enumerate(vals)
    )
    return (
        f"<tr>"
        f'<td class="meta {label_cls}">{esc(label_text)}</td>'
        f'<td class="meta {label_cls}">{side_label}</td>'
        f"{cells}"
        f"</tr>"
    )


def _substitution_summary_html(ignore_substrings):
    """Build HTML lines listing active substitution rules for the summary box."""
    if not ignore_substrings:
        return ""
    lines = ["<p><strong>Substitution rules applied:</strong></p><ul>"]
    for filename, cols in ignore_substrings.items():
        for col, rules in cols.items():
            for find, replacement in rules:
                rep_display = esc(replacement) if replacement else "<em>(empty)</em>"
                lines.append(
                    f'<li>{esc(filename)} / {esc(col)}: '
                    f'<code>{esc(find)}</code> &rarr; {rep_display}</li>'
                )
    lines.append("</ul>")
    return "\n  ".join(lines)


def _nav_html(part, total_parts, base_filename):
    """Build prev/next navigation HTML for split reports. Returns empty string if not split."""
    if total_parts <= 1:
        return ""
    prev_link = (
        f'<a href="{base_filename}_part{part - 1}.html">&larr; Previous</a> &nbsp;|&nbsp; '
        if part > 1 else ""
    )
    next_link = (
        f' &nbsp;|&nbsp; <a href="{base_filename}_part{part + 1}.html">Next &rarr;</a>'
        if part < total_parts else ""
    )
    return f'<div class="nav">{prev_link}Part {part} of {total_parts}{next_link}</div>'


def generate_html(diffs, headers, file1_name, file2_name, case_sensitive=True, ignore_substrings=None,
                  total_rows_compared=0, skip_columns=None, part=1, total_parts=1, base_filename="diff_report",
                  grand_total_changed=None, grand_total_added=None, grand_total_deleted=None,
                  grand_col_diff_counts=None):
    n_changed = sum(1 for d in diffs if d["type"] == "changed")
    n_added   = sum(1 for d in diffs if d["type"] == "added")
    n_deleted = sum(1 for d in diffs if d["type"] == "deleted")

    def _fmt(part_count, grand_total):
        if grand_total is None:
            return str(part_count)
        return f"{part_count}/{grand_total}"

    chg_label = _fmt(n_changed, grand_total_changed)
    add_label = _fmt(n_added,   grand_total_added)
    del_label = _fmt(n_deleted, grand_total_deleted)
    header_strs = [normalize_header(h) for h in headers]

    # Per-column diff counts (changed rows only, excluding skipped columns)
    skip_columns = skip_columns or set()
    col_diff_counts = {h: 0 for h in header_strs if h not in skip_columns}
    for d in diffs:
        if d["type"] == "changed":
            for i in d["changed"]:
                col = header_strs[i]
                if col in col_diff_counts:
                    col_diff_counts[col] += 1

    rows_html = []
    for d in diffs:
        dtype   = d["type"]
        label   = d["label"]
        changed = d["changed"]
        r1, r2  = d["row1"], d["row2"]

        if dtype == "changed":
            rows_html.append(build_data_row(r1, "old", "lbl-old", "OLD", label, changed))
            rows_html.append(build_data_row(r2, "new", "lbl-new", "NEW", label, changed))

        elif dtype == "added":
            cells = "".join(f'<td class="added">{esc(v)}</td>' for v in r2)
            rows_html.append(
                f"<tr>"
                f'<td class="meta lbl-add">{esc(label)}</td>'
                f'<td class="meta lbl-add">ADDED</td>'
                f"{cells}</tr>"
            )

        elif dtype == "deleted":
            cells = "".join(f'<td class="deleted">{esc(v)}</td>' for v in r1)
            rows_html.append(
                f"<tr>"
                f'<td class="meta lbl-del">{esc(label)}</td>'
                f'<td class="meta lbl-del">DELETED</td>'
                f"{cells}</tr>"
            )

    th_cols = "".join(f"<th>{esc(h)}</th>" for h in header_strs)

    if diffs:
        table_html = (
            f"<table>"
            f"<thead><tr><th>Row / Key</th><th>Type</th>{th_cols}</tr></thead>"
            f"<tbody>{''.join(rows_html)}</tbody>"
            f"</table>"
        )
    else:
        table_html = '<p class="ok">&#10003; Files are identical — no differences found.</p>'

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Diff Report</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{
    font-family: Arial, sans-serif;
    font-size: 13px;
    margin: 24px;
    background: #f0f2f5;
    color: #222;
  }}
  h1 {{ margin-bottom: 12px; color: #1e293b; }}
  .summary {{
    background: #fff;
    padding: 14px 20px;
    border-radius: 8px;
    margin-bottom: 22px;
    box-shadow: 0 1px 4px rgba(0,0,0,.12);
  }}
  .summary p {{ margin: 4px 0; }}
  .cnt-chg {{ color: #d97706; font-weight: bold; }}
  .cnt-add {{ color: #16a34a; font-weight: bold; }}
  .cnt-del {{ color: #dc2626; font-weight: bold; }}
  .ok {{ color: #16a34a; font-size: 15px; background: #fff; padding: 16px 20px; border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,.12); }}
  table {{
    width: 100%;
    border-collapse: collapse;
    background: #fff;
    box-shadow: 0 1px 4px rgba(0,0,0,.12);
    border-radius: 8px;
    overflow: hidden;
  }}
  th {{
    background: #1e293b;
    color: #f8fafc;
    padding: 9px 12px;
    text-align: left;
    font-size: 12px;
    white-space: nowrap;
  }}
  td {{
    padding: 6px 12px;
    border-bottom: 1px solid #e5e7eb;
    vertical-align: top;
    word-break: break-word;
    max-width: 280px;
  }}
  td.meta {{
    white-space: nowrap;
    font-size: 11px;
    font-weight: bold;
    min-width: 90px;
  }}
  /* Row type labels */
  td.lbl-old {{ background: #fef2f2; color: #b91c1c; border-right: 3px solid #fca5a5; }}
  td.lbl-new {{ background: #f0fdf4; color: #15803d; border-right: 3px solid #86efac; }}
  td.lbl-add {{ background: #f0fdf4; color: #15803d; border-right: 3px solid #86efac; }}
  td.lbl-del {{ background: #fef2f2; color: #b91c1c; border-right: 3px solid #fca5a5; }}
  /* Row backgrounds */
  td.old     {{ background: #fef2f2; }}
  td.new     {{ background: #f0fdf4; }}
  td.added   {{ background: #dcfce7; }}
  td.deleted {{ background: #fee2e2; }}
  /* Highlighted (changed) cell */
  td.cell-hl {{ background: #fef08a !important; font-weight: bold; outline: 1px solid #ca8a04; }}
  tr:last-child td {{ border-bottom: none; }}
  /* Separator between diff groups */
  tr.sep td {{ height: 6px; background: #f0f2f5; border: none; padding: 0; }}
  .nav {{ margin-bottom: 16px; font-size: 13px; }}
  .nav a {{ color: #1e293b; text-decoration: none; font-weight: bold; }}
  .nav a:hover {{ text-decoration: underline; }}
</style>
</head>
<body>
<h1>Diff Report</h1>
<div class="summary">
  <p><strong>File 1:</strong> {esc(file1_name)}</p>
  <p><strong>File 2:</strong> {esc(file2_name)}</p>
  <p><strong>Comparison mode:</strong> {"Case-sensitive" if case_sensitive else "Case-insensitive"}</p>
  <p><strong>Total rows compared:</strong> {total_rows_compared}</p>
  {_substitution_summary_html(ignore_substrings)}
  <br>
  <p class="cnt-chg">&#9679; Changed rows: {chg_label}</p>
  <p class="cnt-add">&#9679; Added rows &nbsp;(only in File 2): {add_label}</p>
  <p class="cnt-del">&#9679; Deleted rows (only in File 1): {del_label}</p>
  <br>
  <p><strong>Changed cells per column:</strong></p>
  <ul>
  {"".join(f'<li><strong>{esc(col)}:</strong> {_fmt(count, grand_col_diff_counts.get(col) if grand_col_diff_counts else None)} row(s) differ</li>' for col, count in col_diff_counts.items() if count > 0 or (grand_col_diff_counts and grand_col_diff_counts.get(col, 0) > 0))}
  </ul>
</div>
{_nav_html(part, total_parts, base_filename)}
{table_html}
{_nav_html(part, total_parts, base_filename)}
</body>
</html>"""


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=== Diff Tool ===")

    # 1. Resolve the two files
    file1, file2 = get_files()
    print(f"\nComparing:")
    print(f"  File 1: {os.path.basename(file1)}")
    print(f"  File 2: {os.path.basename(file2)}")

    # 2. Read both files (first sheet, ignore empty cols/trailing rows)
    print("\nReading files...")
    headers1, rows1 = read_sheet(file1)
    headers2, rows2 = read_sheet(file2)
    print(f"  File 1: {len(rows1)} data rows, {len(headers1)} columns")
    print(f"  File 2: {len(rows2)} data rows, {len(headers2)} columns")

    # 3. Validate same column structure
    print("Checking column format...")
    check_format(headers1, headers2)
    print("  OK — columns match.")

    # 4. Ask export format (before all other prompts)
    export_format = ask_export_format()

    # 5. Load aliases and substitution rules
    aliases = load_aliases()
    ignore_substrings = load_ignore_substrings()
    if aliases:
        header_strs = [normalize_header(h) for h in headers1]
        for col in list(aliases.keys()):
            if col not in header_strs:
                print(f"  Warning: alias column '{col}' not found in headers, skipping.")
                del aliases[col]

    # 6. Ask about unique key column
    has_key, key_col = ask_unique_key(headers1)

    # 7. Ask about comparison mode
    case_sensitive = ask_case_sensitive()

    # 8. Ask about columns to skip
    skip_columns = ask_skip_columns(headers1)

    # 9. Without a key column, row counts must match
    if not has_key and len(rows1) != len(rows2):
        raise ValueError(
            f"Row count mismatch (File1={len(rows1)}, File2={len(rows2)}).\n"
            f"Tip: use a unique key column to support added/deleted row detection."
        )

    # 10. Compare
    print("\nComparing rows...")
    f1_name = os.path.basename(file1)
    f2_name = os.path.basename(file2)
    if has_key:
        diffs = compare_by_key(headers1, rows1, rows2, key_col, case_sensitive, aliases,
                               skip_columns, ignore_substrings, f1_name, f2_name)
    else:
        diffs = compare_by_position(headers1, rows1, rows2, case_sensitive, aliases,
                                    skip_columns, ignore_substrings, f1_name, f2_name)

    # 11. Ask about extra post-substitution report (and split only for HTML)
    want_extra = ask_extra_report()
    should_split, rows_per_file = ask_split_report() if export_format == "html" else (False, None)

    # 12. Build shared kwargs for report generators
    shared_kwargs = dict(
        file1_name=f1_name,
        file2_name=f2_name,
        case_sensitive=case_sensitive,
        ignore_substrings=ignore_substrings,
        total_rows_compared=max(len(rows1), len(rows2)),
        skip_columns=skip_columns,
    )

    # Helper: resolve output path and write content
    def write_report(content, filename, binary=False):
        mode = "wb" if binary else "w"
        encoding = None if binary else "utf-8"
        path = os.path.join(os.getcwd(), filename)
        with open(path, mode, encoding=encoding) as fh:
            if binary:
                content.save(fh)
            else:
                fh.write(content)
        return path

    def write_report_to_dir(content, directory, filename, binary=False):
        mode = "wb" if binary else "w"
        encoding = None if binary else "utf-8"
        path = os.path.join(directory, filename)
        with open(path, mode, encoding=encoding) as fh:
            if binary:
                content.save(fh)
            else:
                fh.write(content)
        return path

    # 13. Generate main report
    out_path = None
    if export_format == "html":
        if should_split:
            reports_dir = os.path.join(os.getcwd(), "reports")
            if os.path.exists(reports_dir):
                shutil.rmtree(reports_dir)
            os.makedirs(reports_dir)
            grand_changed = sum(1 for d in diffs if d["type"] == "changed")
            grand_added   = sum(1 for d in diffs if d["type"] == "added")
            grand_deleted = sum(1 for d in diffs if d["type"] == "deleted")
            _hstrs = [normalize_header(h) for h in headers1]
            _skip  = skip_columns or set()
            grand_col_counts = {h: 0 for h in _hstrs if h not in _skip}
            for d in diffs:
                if d["type"] == "changed":
                    for ci in d["changed"]:
                        col = _hstrs[ci]
                        if col in grand_col_counts:
                            grand_col_counts[col] += 1
            chunks = [diffs[i:i + rows_per_file] for i in range(0, max(len(diffs), 1), rows_per_file)]
            total_parts = len(chunks)
            for i, chunk in enumerate(chunks, 1):
                html = generate_html(chunk, headers1, part=i, total_parts=total_parts,
                                     base_filename="diff_report",
                                     grand_total_changed=grand_changed,
                                     grand_total_added=grand_added,
                                     grand_total_deleted=grand_deleted,
                                     grand_col_diff_counts=grand_col_counts,
                                     **shared_kwargs)
                write_report_to_dir(html, reports_dir, f"diff_report_part{i}.html")
            print(f"Report split into {total_parts} file(s) in 'reports' folder.")
            out_path = os.path.join(reports_dir, "diff_report_part1.html")
        else:
            html = generate_html(diffs, headers1, **shared_kwargs)
            out_path = write_report(html, "diff_report.html")
    elif export_format == "csv":
        content = generate_csv_report(diffs, headers1, **shared_kwargs)
        out_path = write_report(content, "diff_report.csv")
    elif export_format == "xlsx":
        wb = generate_xlsx_report(diffs, headers1, **shared_kwargs)
        out_path = write_report(wb, "diff_report.xlsx", binary=True)

    # 14. Generate extra post-substitution report if requested
    if want_extra:
        header_strs = [normalize_header(h) for h in headers1]
        col_sub1 = _build_col_subs(header_strs, ignore_substrings, f1_name)
        col_sub2 = _build_col_subs(header_strs, ignore_substrings, f2_name)
        changed_diffs = [d for d in diffs if d["type"] == "changed"]
        if not changed_diffs:
            print("No changed rows found — extra report not generated.")
        elif export_format == "html" and should_split:
            all_diff_col_indices = sorted({j for d in changed_diffs for j in d["changed"]})
            chunks = [changed_diffs[i:i + rows_per_file] for i in range(0, max(len(changed_diffs), 1), rows_per_file)]
            total_parts = len(chunks)
            for i, chunk in enumerate(chunks, 1):
                extra_html = generate_extra_html(chunk, headers1, f1_name, f2_name, col_sub1, col_sub2,
                                                 part=i, total_parts=total_parts,
                                                 base_filename="diff_report_substituted",
                                                 diff_col_indices=all_diff_col_indices)
                if extra_html:
                    write_report_to_dir(extra_html, reports_dir, f"diff_report_substituted_part{i}.html")
            print(f"Extra report split into {total_parts} file(s) in 'reports' folder.")
        elif export_format == "html":
            extra_html = generate_extra_html(changed_diffs, headers1, f1_name, f2_name, col_sub1, col_sub2)
            if extra_html:
                extra_path = write_report(extra_html, "diff_report_substituted.html")
                print(f"Extra report saved to: {extra_path}")
        elif export_format == "csv":
            content = generate_extra_csv(changed_diffs, headers1, f1_name, f2_name, col_sub1, col_sub2)
            if content:
                extra_path = write_report(content, "diff_report_substituted.csv")
                print(f"Extra report saved to: {extra_path}")
        elif export_format == "xlsx":
            wb = generate_extra_xlsx(changed_diffs, headers1, f1_name, f2_name, col_sub1, col_sub2)
            if wb:
                extra_path = write_report(wb, "diff_report_substituted.xlsx", binary=True)
                print(f"Extra report saved to: {extra_path}")

    # 15. Print summary
    n_changed = sum(1 for d in diffs if d["type"] == "changed")
    n_added   = sum(1 for d in diffs if d["type"] == "added")
    n_deleted = sum(1 for d in diffs if d["type"] == "deleted")
    print(f"\nResults:")
    print(f"  Changed : {n_changed}")
    print(f"  Added   : {n_added}")
    print(f"  Deleted : {n_deleted}")
    print(f"\nReport saved to: {out_path}")


if __name__ == "__main__":
    main()
